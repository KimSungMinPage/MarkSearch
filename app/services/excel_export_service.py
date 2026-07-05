from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
from html import escape
from app.models import Company, Factory
from app.services.statistics_service import StatisticsService
try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
except Exception:  # Linux CI may be offline; Windows build installs openpyxl from requirements.
    Workbook = None
class ExcelExportService:
    def export(self, path: str | Path, companies: list[Company], factories: list[Factory]) -> None:
        if Workbook is None: return self._export_minimal_xlsx(Path(path), companies, factories)
        wb=Workbook(); ws=wb.active; ws.title="수집정보"; ws.append(["항목","값"]); ws.append(["API 서비스명","의약품 등 업체허가현황"]); ws.append(["업체 목록 총 건수",len(companies)]); ws.append(["제조소 상세 총 건수",len(factories)])
        cws=wb.create_sheet("업체목록"); c_headers=["번호","업종","업체명","정규화업체명","대표자명","업체허가일자","업체허가일자원본","업체허가번호","공장주소","시도","시군구","중복여부","중복그룹번호"]; cws.append(c_headers)
        for i,r in enumerate(companies,1): cws.append([i,r.induty,r.entrps,r.normalized_entrps,r.rprsntv,r.prmisn_date_display,r.prmisn_dt,str(r.prmisn_no),r.adres,r.sido,r.sigungu,"Y" if r.is_duplicate else "N",r.duplicate_group_no])
        fws=wb.create_sheet("제조소상세"); f_headers=["번호","업체명","정규화업체명","업종","업체허가번호","대표자명","업체허가일자","업체허가일자원본","공장일련번호","공장명","공장전화번호","공장주소1","공장주소2","통합공장주소","시도","시군구","중복여부","중복그룹번호"]; fws.append(f_headers)
        for i,r in enumerate(factories,1): fws.append([i,r.entrps,r.normalized_entrps,r.induty,str(r.prmisn_no),r.rprsntv,r.prmisn_date_display,r.prmisn_dt,r.sn,r.fctry_nm,str(r.telno),r.fctry_addr1,r.fctry_addr2,r.combined_address,r.sido,r.sigungu,"Y" if r.is_duplicate else "N",r.duplicate_group_no])
        stat=StatisticsService(); iws=wb.create_sheet("업종별통계"); iws.append(["업종","업체수","제조소수"]); [iws.append(list(row)) for row in stat.industry_counts(companies,factories)]
        rws=wb.create_sheet("지역별통계"); rws.append(["시도","업체수","제조소수"]); [rws.append(list(row)) for row in stat.region_counts(companies,factories)]
        wb.create_sheet("오류및누락").append(["유형","내용"])
        for sheet in wb.worksheets: self._format(sheet)
        wb.save(path)
    def _format(self, ws):
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions; fill=PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]: cell.font=Font(bold=True); cell.fill=fill
        for col in ws.columns:
            max_len=max(len(str(c.value or "")) for c in col); ws.column_dimensions[col[0].column_letter].width=min(max(max_len+2,10),60)
            for c in col: c.alignment=Alignment(wrap_text=True)
    def _export_minimal_xlsx(self, path: Path, companies: list[Company], factories: list[Factory]) -> None:
        sheets = {
            "수집정보":[["항목","값"],["API 서비스명","의약품 등 업체허가현황"],["업체 목록 총 건수",len(companies)],["제조소 상세 총 건수",len(factories)]],
            "업체목록":[["번호","업체명","업체허가번호"]]+[[i,c.entrps,c.prmisn_no] for i,c in enumerate(companies,1)],
            "제조소상세":[["번호","업체명","업체허가번호","공장전화번호"]]+[[i,f.entrps,f.prmisn_no,f.telno] for i,f in enumerate(factories,1)],
            "업종별통계":[["업종","업체수","제조소수"]],"지역별통계":[["시도","업체수","제조소수"]],"오류및누락":[["유형","내용"]],
        }
        with ZipFile(path,"w",ZIP_DEFLATED) as z:
            z.writestr("[Content_Types].xml", '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>' + ''.join(f'<Override PartName="/xl/worksheets/sheet{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>' for i in range(1,len(sheets)+1)) + '</Types>')
            z.writestr("_rels/.rels", '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>')
            z.writestr("xl/_rels/workbook.xml.rels", '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">' + ''.join(f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i}.xml"/>' for i in range(1,len(sheets)+1)) + '</Relationships>')
            z.writestr("xl/workbook.xml", '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets>' + ''.join(f'<sheet name="{escape(name)}" sheetId="{i}" r:id="rId{i}"/>' for i,name in enumerate(sheets,1)) + '</sheets></workbook>')
            for i, rows in enumerate(sheets.values(),1):
                body=''.join('<row>' + ''.join(f'<c t="inlineStr"><is><t>{escape(str(v))}</t></is></c>' for v in row) + '</row>' for row in rows)
                z.writestr(f"xl/worksheets/sheet{i}.xml", f'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>{body}</sheetData></worksheet>')
